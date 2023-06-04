import openpyxl
import numpy as np

wb = openpyxl.Workbook()
thrbin = ["000", "001", "010", "011", "100", "101", "110", "111"] #three-bit binary
pi = ["001", "100", "111", "010", "101", "000", "011", "110"] #pi(x) = 3x+1 mod 8 

resA = np.zeros([8, 8], dtype=int)

#for excel
sheet = wb.worksheets[0]
sheet.cell(row=1, column=1, value="a")
sheet.cell(row=1, column=2, value="x")
sheet.cell(row=1, column=3, value="result")
sheet.cell(row=1, column=4, value="Sset")

#calculation process
m = 2
for a in range(8):
    sheet.cell(row=m, column=1, value=a)
    for x in range(8):
        sheet.cell(row=m, column=2, value=x)
        res = int(thrbin[a][0]) * int(pi[x][0]) ^ int(thrbin[a][1]) * int(pi[x][1]) ^ int(thrbin[a][2]) * int(pi[x][2])
        resA[a][x] = res
        sheet.cell(row=m, column=3, value=res)
        m += 1

for a in range(8):
    Sset = ""
    for n in range(8):
        if resA[a][0] == resA[a][n]:
            Sset += str(n) + ","
    if Sset:
        Sset = Sset[:-1]  # remove the last comma
    sheet.cell(row=a + 2, column=4, value=Sset)

wb.save("simon_find_s.xlsx")

